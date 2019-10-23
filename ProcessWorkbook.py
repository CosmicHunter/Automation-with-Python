import  openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):
        workBook_object = xl.load_workbook(filename) # This methods loads the excel sheet and loads the workbook object

        # now in the excel we only have one sheet now to excess that
        sheet = workBook_object['Sheet1']
        # to access a particular cell of the sheet
        cell = sheet['a1']   # this means A column and row 1 (cordinates of a specific cell)
        cell = sheet.cell(1,1) # this also does the same thing in this we call the cell method of the sheet object
        # to access the value of the cell we use cell.value
        print(cell.value)
        print('This is for the new branch commit to make a pull request in this repository')

        #now we will iterate over each row and multiply price in each row by 0.9
        # now sheet.max_row gets the maximum number of rows in the sheet
        # so we will iterate over the row

        for ptr in range(2,sheet.max_row+1):
            cell = sheet.cell(ptr,3)
            #print(cell.value)
            # now each price we will multiply by 0.9
            corrected_price = cell.value * 0.9
            # now we will add these prices to new column
            corrected_price_cell = sheet.cell(ptr,4)
            corrected_price_cell.value = corrected_price

        values = Reference(sheet,
                           min_row=2,
                           max_row=sheet.max_row,
                           min_col=4,
                           max_col=4)


        # This will allow us to access the elements from the sheet from row 2 to 4 column
        # column = 4


        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'E2')
        # this will add a bar chart starting from
        # e2
        # After doing modifications we need to save the workbooks
        # We can do that in same sheet or a new sheet
        workBook_object.save('transactions2.xlsx')

# to modify the file call the function process_workbook
